import os
import pandas as pd
import logging
from datetime import datetime
from helpers.db_conn import get_db
from openpyxl import load_workbook as pyxl_load_workbook
from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font


logger = logging.getLogger(__name__)

"""
This is for generating upload files for

RV
SL
CR
Transfer

excel files for a date based off CDT, CDP, TransactionFile

Replicates the SQL Server 'Daily Rerun' job steps 8-29 entirely in-memory.
"""

#take in full nework path with escapes e.g \\\\tutpub3\\VOL2\\TestFiles\\Manual_Reruns\\rerun_resources
def proccess_transfile(transfile_path: str) -> pd.DataFrame:
    transfile_df = pd.read_csv(transfile_path, sep='\t', header = None, dtype = {
        0 : str,
        1 : str,
        2 : str,
        3 : str,
        4 : str,
        5 : str,
        6 : str,
        7 : str,
        8 : str,
        9 : str,
        10 : str,
        11 : str,
        12 : int,
        13 : float,
        14 : float,
        15 : float,
        16 : str,
        17 : str,
        18 : str,
        19 : str,
        20 : str,
        21 : str,
    })

    transfile_df.columns = [
        "Ordnum","Otype","Otypesra",
        "Ponumber","Billto","Billtoname",
        "Bcntry","Shipto","Shipname",
        "ISBN","Title","Client",
        "Qty","Ext","Price","Discount"
        ,"Currenttyp","RettyP","Linekey",
        "Ingwhs","St_name","Pdate"
    ]

    return transfile_df

def procces_cdt_file(cdt_path : str) -> pd.DataFrame:
    cdt_df = pd.read_csv(cdt_path, header = None, dtype  = {
        0 : str,
        1 : str,
        2 : str,
        3 : str,
        4 : str,
        5 : int,
        6 : str,
        7 : int,
        8 : str,
        9 : str,
        10 : str,
        11 : str,
        12 : str
    })

    cdt_df.columns = [
        "TutNum","FileDate","WHS",
        "ISBN10","UPS","EAN",
        "Transcode","Qty","Each",
        "Dispcode","Linenum","Acttype",
        "FromLoc"
    ]

    cdt_df['ToLoc'] = None
    cdt_df['PONum'] = None
    cdt_df['Qtyreq'] = None

    def whs_updates(x: str) -> str:
        if x.startswith('631760'):
            return "IPS"
        elif x == "6318681":
            return "DAMAGE"
        else:
            return "ING"

    cdt_df['WHS'] = cdt_df['WHS'].apply(lambda x : whs_updates(x))

    #DELETE FROM ips.dbo.IPS_INV WHERE Acttype = 'SS' OR Acttype = 'IM' OR Acttype ='RT'
    cdt_df = cdt_df[~cdt_df["Acttype"].isin(['SS','IM','RT'])]

    # UPDATE ips.dbo.[IPS_INV] SET FromLoc = 'DAMAGE', ToLoc = WHS WHERE Acttype = '??' OR Acttype = 'HS' OR Acttype ='HD';

    boolean_mask_1 = cdt_df["Acttype"].isin(['??','HS','HD'])
    cdt_df.loc[boolean_mask_1, 'FromLoc'] = "DAMAGE"
    cdt_df.loc[boolean_mask_1, 'ToLoc'] = cdt_df.loc[boolean_mask_1, 'WHS']

    #UPDATE ips.dbo.[IPS_INV] SET FromLoc = WHS , ToLoc = 'DAMAGE' WHERE Acttype = 'DT';

    boolean_mask_2 = cdt_df["Acttype"] == 'DT'
    cdt_df.loc[boolean_mask_2, 'FromLoc'] = cdt_df.loc[boolean_mask_2, 'WHS']
    cdt_df.loc[boolean_mask_2, 'ToLoc'] = "DAMAGE"

    #UPDATE ips.dbo.[IPS_INV] SET FromLoc = WHS , ToLoc = 'ING' WHERE Acttype ='TC' OR Acttype = 'TD' OR Acttype = 'TE' OR Acttype = 'TN' OR Acttype = 'TH';
    boolean_mask_3 = cdt_df["Acttype"].isin(['TC','TD','TE','TN','TH'])
    cdt_df.loc[boolean_mask_3, 'FromLoc'] = cdt_df.loc[boolean_mask_3, 'WHS']
    cdt_df.loc[boolean_mask_3, 'ToLoc'] = 'ING'

    #UPDATE ips.dbo.[IPS_INV] SET Qty = Qty * -1 Where Fromloc <> '' and QTY < 0;
    boolean_mask_4 = cdt_df['FromLoc'].notna() & (cdt_df['FromLoc'] != '') & (cdt_df['Qty'] < 0)
    cdt_df.loc[boolean_mask_4,'Qty'] = cdt_df.loc[boolean_mask_4,'Qty'] * -1

    #UPDATE ips.dbo.[IPS_INV] SET Qtyreq = Qty Where Fromloc <> '';
    boolean_mask_5 = cdt_df['FromLoc'].notna() & (cdt_df['FromLoc'] != '')
    cdt_df.loc[boolean_mask_5,'Qtyreq'] = cdt_df.loc[boolean_mask_5,'Qty']

    #UPDATE ips.dbo.[IPS_INV] SET PONum = 'IPS_TRANS_' + REPLACE(CONVERT(nchar(50), GETDATE(),1), '/', '') Where Fromloc <> '';
    date_str = datetime.now().strftime('%m/%d/%y').replace('/', '')
    boolean_mask_ponum = cdt_df['FromLoc'].notna() & (cdt_df['FromLoc'] != '')
    cdt_df.loc[boolean_mask_ponum, 'PONum'] = f'IPS_TRANS_{date_str}'

    return cdt_df


def proccess_cdp_file(cdp_path: str) -> pd.DataFrame:
    """
    Reads the LOCKED.CDP file (comma-delimited, 13 columns) into a DataFrame.
    Mirrors the column structure of the dbo.LOCKED table from Locked_Import.dtsx.
    """
    cdp_df = pd.read_csv(cdp_path, header=None, dtype={
        0: str, 1: str, 2: str, 3: str, 4: str, 5: str, 6: str,
        7: int,
        8: str, 9: str, 10: str, 11: str, 12: str
    })
    cdp_df.columns = [
        'F1', 'Fdate', 'San', 'ISBN10', 'F5', 'ISBN',
        'Invcode', 'QTY',
        'column9', 'column10', 'column11', 'column12', 'column13'
    ]
    return cdp_df


#take in full nework path with escapes e.g \\\\tutpub3\\VOL2\\TestFiles\\Manual_Reruns\\rerun_resources
def proccess_daily_files_rerun(CDT_path: str, CDP_path: str, transfile_path: str, output_path: str) -> None:
    os.makedirs(output_path, exist_ok=True)

    # ─── Phase 1: CDT → IPS_INV → Transfer outputs (steps 8-10) ─────────────
    logger.info("Processing CDT file")
    inv_df = procces_cdt_file(CDT_path)

    # Transfer.csv — matches Transfer output.dtsx: fromloc starts with 'I'
    transfer_mask = inv_df['FromLoc'].notna() & inv_df['FromLoc'].str[:1].eq('I')
    transfer_raw = inv_df.loc[transfer_mask, ['Linenum', 'EAN', 'FromLoc', 'ToLoc', 'Qty', 'PONum', 'Qtyreq']].copy()
    transfer_raw.columns = ['linenum', 'Ean', 'Fromloc', 'Toloc', 'Qty', 'Ponum', 'Qtyreq']

    transfer_csv_path = os.path.join(output_path, 'Transfer.csv')
    transfer_raw.to_csv(transfer_csv_path, index=False)
    logger.info(f"Transfer.csv written: {len(transfer_raw)} rows → {transfer_csv_path}")

    # TRANSFER_SAGE_UPLOAD.xlsx — Sage300 import format
    transfer_sage = transfer_raw.copy()
    transfer_sage.columns = ['LINENUM', 'EAN', 'FROMLOC', 'TOLOC', 'QTY', 'PONUM', 'QTYREQ']
    transfer_sage_path = os.path.join(output_path, 'TRANSFER_SAGE_UPLOAD.xlsx')
    for col in transfer_sage.columns:
        transfer_sage[col] = transfer_sage[col].astype(str)
    with pd.ExcelWriter(transfer_sage_path) as f:
        transfer_sage.to_excel(f, sheet_name='Transfer', index=False)
    wb: Workbook = pyxl_load_workbook(transfer_sage_path)
    for sheet in ['Transfer']:
        ws: Worksheet = wb[sheet]
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].number_format = '@'
        for cell in ws[1]:
            cell.font = Font(bold=False, underline='none')
        cell_range = f"'{sheet}'!$A$1:${get_column_letter(ws.max_column)}${ws.max_row}"
        wb.defined_names[sheet] = DefinedName(name=sheet, attr_text=cell_range)
    wb.save(transfer_sage_path)
    wb.close()
    logger.info(f"TRANSFER_SAGE_UPLOAD.xlsx written → {transfer_sage_path}")

    # ─── Phase 2: TransactionFile → IPS_DAILY (step 11) ──────────────────────
    logger.info("Processing TransactionFile")
    daily_df = proccess_transfile(transfile_path)

    # Generate Line_num and Order_id — mirrors FIX.py logic
    orddict = {}
    ordIddict = {}
    id_start = 6300
    line_nums = []
    order_ids = []
    for ordnum in daily_df['Ordnum']:
        if ordnum in orddict:
            orddict[ordnum] += 1
        else:
            orddict[ordnum] = 1
        line_nums.append(orddict[ordnum])
        if ordnum not in ordIddict:
            ordIddict[ordnum] = str(id_start)
            id_start += 1
        order_ids.append(ordIddict[ordnum])
    daily_df['Line_num'] = line_nums
    daily_df['Order_id'] = order_ids

    # ─── Step 15: IPS_DAILY Queries Part 1 ───────────────────────────────────
    daily_df = daily_df[daily_df['Qty'] != 0]
    daily_df = daily_df[~daily_df['ISBN'].isin(['0', ''])].copy()
    daily_df = daily_df[~daily_df['Billto'].isin(['000799074', '000647955'])]
    daily_df = daily_df[~daily_df['ISBN'].str[:8].eq('97814629')]

    # UPDATE SET Billto = Shipto Where Billto = '000808073'
    mask_808 = daily_df['Billto'] == '000808073'
    daily_df.loc[mask_808, 'Billto'] = daily_df.loc[mask_808, 'Shipto']

    # Crossref lookup: remap Billto → Ssacct, set Crossref = 'X'
    daily_df['Crossref'] = None
    try:
        with get_db() as conn:
            crossref_df = pd.read_sql('SELECT Billto, Ssacct FROM IPS.dbo.crossref', con=conn)
        crossref_map = dict(zip(crossref_df['Billto'], crossref_df['Ssacct']))
        xref_mask = daily_df['Billto'].isin(crossref_map.keys())
        daily_df.loc[xref_mask, 'Crossref'] = 'X'
        daily_df.loc[xref_mask, 'Billto'] = daily_df.loc[xref_mask, 'Billto'].map(crossref_map)
        logger.info(f"Crossref applied: {xref_mask.sum()} rows remapped")
    except Exception as e:
        logger.warning(f"Crossref lookup failed — Billto will not be remapped: {e}")

    # Ensure float for numeric comparisons
    daily_df['Discount'] = daily_df['Discount'].astype(float)
    daily_df['Price'] = daily_df['Price'].astype(float)
    daily_df['Ext'] = daily_df['Ext'].astype(float)

    # UPDATE SET Otype = 'Return' where Rettyp matches return codes, Price=0, or Ext<0
    return_rettyps = {'20','50','3501','2008','2020','3520','3509','3508','2509','2501','2009','2002','2001','2018'}
    return_mask = (
        daily_df['RettyP'].isin(return_rettyps) |
        (daily_df['Price'] == 0) |
        (daily_df['Ext'] < 0)
    )
    daily_df.loc[return_mask, 'Otype'] = 'Return'

    # UPDATE SET Price = Price * -1 WHERE Price < 0
    daily_df.loc[daily_df['Price'] < 0, 'Price'] = daily_df.loc[daily_df['Price'] < 0, 'Price'] * -1

    # UPDATE SET Discount = 100.0 WHERE Ext = 0.00
    daily_df.loc[daily_df['Ext'] == 0.00, 'Discount'] = 100.0

    # UPDATE SET Whs from Ingwhs
    daily_df['Whs'] = 'ING'
    daily_df.loc[daily_df['Ingwhs'] == 'HH', 'Whs'] = 'IPS'
    damage_rettyps = {'20','3501','3520','50','2008','2020','3509','2509','2501','2009','2002','2001'}
    daily_df.loc[daily_df['RettyP'].isin(damage_rettyps), 'Whs'] = 'DAMAGE'

    # ─── Step 17: IPS_DAILY Queries Part 2 ───────────────────────────────────
    # UPDATE SET Ordnum = Ordnum + 'I'
    daily_df['Ordnum'] = daily_df['Ordnum'] + 'I'

    # DELETE WHERE SUBSTRING(Otypesra,1,1) = 'S' or 'R'
    daily_df = daily_df[~daily_df['Otypesra'].str[:1].isin(['S', 'R'])].copy()

    # Derived field updates
    daily_df['Rep_inv'] = daily_df['Ordnum']
    daily_df['Repqty'] = daily_df['Qty']
    daily_df['Review'] = 'REVIEW'
    daily_df['Post'] = 'FALSE'

    # Traninfo: 'TRANSFILE_mmddyy' for non-returns, 'TRANSFILEREV_mmddyy' for reviews (Discount=100)
    pdate_series = pd.to_datetime(daily_df['Pdate'], errors='coerce')
    daily_df['Traninfo'] = None
    non_return_mask = daily_df['Otype'] != 'Return'
    daily_df.loc[non_return_mask, 'Traninfo'] = 'TRANSFILE_' + pdate_series[non_return_mask].dt.strftime('%m%d%y')
    review_mask = daily_df['Discount'] == 100.0
    daily_df.loc[review_mask, 'Traninfo'] = 'TRANSFILEREV_' + pdate_series[review_mask].dt.strftime('%m%d%y')

    logger.info(f"IPS_DAILY in-memory: {len(daily_df)} rows after all transforms")

    # ─── Steps 18-19: Export RV ───────────────────────────────────────────────
    rv_df = daily_df[daily_df['Discount'] == 100.0].copy()

    rv_header = rv_df[rv_df['Line_num'] == 1][
        ['Order_id', 'Ordnum', 'Billto', 'Ponumber', 'Pdate', 'Rep_inv', 'Traninfo', 'Post']
    ].copy()
    rv_header.columns = ['ORDUNIQ', 'ORDNUMBER', 'CUSTOMER', 'PONUMBER', 'ORDDATE', 'DESC', 'COMMENT', 'POSTINV']
    rv_header['ORDDATE'] = pd.to_datetime(rv_header['ORDDATE'], errors='coerce').apply(
        lambda x: f"{x.month}/{x.day}/{x.year}" if pd.notna(x) else None
    )

    rv_detail = rv_df[['Order_id', 'Line_num', 'ISBN', 'Review', 'Whs', 'Qty', 'Price', 'Discount', 'Repqty']].copy()
    rv_detail.columns = ['ORDUNIQ', 'LINENUM', 'ITEM', 'REVIEW', 'LOCATION', 'QTYORDERED', 'PRIUNTPRC', 'DISCPER', 'QTYSHIPPED']

    rv_path = os.path.join(output_path, 'RV_SAGE_UPLOAD.xlsx')
    for col in rv_header.columns:
        rv_header[col] = rv_header[col].astype(str)
    for col in rv_detail.columns:
        rv_detail[col] = rv_detail[col].astype(str)
    with pd.ExcelWriter(rv_path) as f:
        rv_header.to_excel(f, sheet_name='RV_Header', index=False)
        rv_detail.to_excel(f, sheet_name='RV_Detail', index=False)
    wb: Workbook = pyxl_load_workbook(rv_path)
    for sheet in ['RV_Header', 'RV_Detail']:
        ws: Worksheet = wb[sheet]
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].number_format = '@'
        for cell in ws[1]:
            cell.font = Font(bold=False, underline='none')
        cell_range = f"'{sheet}'!$A$1:${get_column_letter(ws.max_column)}${ws.max_row}"
        wb.defined_names[sheet] = DefinedName(name=sheet, attr_text=cell_range)
    wb.save(rv_path)
    wb.close()
    logger.info(f"RV written: {len(rv_header)} header, {len(rv_detail)} detail → {rv_path}")

    # ─── Step 20: Delete reviews from working set 
    daily_df = daily_df[daily_df['Discount'] != 100.0].copy()

    #  Steps 21-22: Export CR 
    cr_df = daily_df[daily_df['Otype'] == 'Return'].copy()

    cr_header = cr_df[cr_df['Line_num'] == 1][
        ['Order_id', 'Ordnum', 'Billto', 'Ponumber', 'Pdate']
    ].copy()
    cr_header.columns = ['CRDUNIQ', 'ORDNUMBER', 'CUSTOMER', 'PONUMBER', 'ORDDATE']
    cr_header['ORDDATE'] = pd.to_datetime(cr_header['ORDDATE'], errors='coerce').apply(
        lambda x: f"{x.month}/{x.day}/{x.year}" if pd.notna(x) else None
    )

    cr_detail = cr_df[['Order_id', 'Line_num', 'ISBN', 'Qty', 'Price', 'Discount']].copy()
    cr_detail.columns = ['CRDUNIQ', 'LINENUM', 'ITEM', 'QTYRETURN', 'PRIUNTPRC', 'DISCPER']
    cr_detail.insert(3, 'LOCATION', 'DAMAGE')

    cr_path = os.path.join(output_path, 'CR_SAGE_UPLOAD.xlsx')
    for col in cr_header.columns:
        cr_header[col] = cr_header[col].astype(str)
    for col in cr_detail.columns:
        cr_detail[col] = cr_detail[col].astype(str)
    with pd.ExcelWriter(cr_path) as f:
        cr_header.to_excel(f, sheet_name='Credit_Debit_Notes', index=False)
        cr_detail.to_excel(f, sheet_name='Credit_Debit_Detail', index=False)
    wb: Workbook = pyxl_load_workbook(cr_path)
    for sheet in ['Credit_Debit_Notes', 'Credit_Debit_Detail']:
        ws: Worksheet = wb[sheet]
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].number_format = '@'
        for cell in ws[1]:
            cell.font = Font(bold=False, underline='none')
        cell_range = f"'{sheet}'!$A$1:${get_column_letter(ws.max_column)}${ws.max_row}"
        wb.defined_names[sheet] = DefinedName(name=sheet, attr_text=cell_range)
    wb.save(cr_path)
    wb.close()
    logger.info(f"CR written: {len(cr_header)} header, {len(cr_detail)} detail → {cr_path}")

    # ─── Steps 23-24: Export SLs
    sl_df = daily_df[daily_df['Otype'] != 'Return'].copy()

    sl_header = sl_df[sl_df['Line_num'] == 1][
        ['Order_id', 'Ordnum', 'Billto', 'Ponumber', 'Pdate', 'Rep_inv', 'Traninfo', 'Post']
    ].copy()
    sl_header.columns = ['ORDUNIQ', 'ORDNUMBER', 'CUSTOMER', 'PONUMBER', 'ORDDATE', 'DESC', 'COMMENT', 'POSTINV']
    sl_header['ORDDATE'] = pd.to_datetime(sl_header['ORDDATE'], errors='coerce').apply(
        lambda x: f"{x.month}/{x.day}/{x.year}" if pd.notna(x) else None
    )

    sl_detail = sl_df[['Order_id', 'Line_num', 'ISBN', 'Whs', 'Qty', 'Price', 'Discount', 'Repqty']].copy()
    sl_detail.columns = ['ORDUNIQ', 'LINENUM', 'ITEM', 'LOCATION', 'QTYORDERED', 'PRIUNTPRC', 'DISCPER', 'QTYSHIPPED']

    sl_path = os.path.join(output_path, 'SL_SAGE_UPLOAD.xlsx')
    for col in sl_header.columns:
        sl_header[col] = sl_header[col].astype(str)
    for col in sl_detail.columns:
        sl_detail[col] = sl_detail[col].astype(str)

    with pd.ExcelWriter(sl_path) as f:
        sl_header.to_excel(f, sheet_name='Orders', index=False)
        sl_detail.to_excel(f, sheet_name='Order_Details', index=False)
    wb: Workbook = pyxl_load_workbook(sl_path)
    for sheet in ['Orders', 'Order_Details']:
        ws: Worksheet = wb[sheet]
        for col_idx in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col_idx)].number_format = '@'
        for cell in ws[1]:
            cell.font = Font(bold=False, underline='none')
        cell_range = f"'{sheet}'!$A$1:${get_column_letter(ws.max_column)}${ws.max_row}"
        wb.defined_names[sheet] = DefinedName(name=sheet, attr_text=cell_range)
    wb.save(sl_path)
    wb.close()
    logger.info(f"SL written: {len(sl_header)} header, {len(sl_detail)} detail → {sl_path}")

    # ─── Step 25: ING_Transfers.csv ───────────────────────────────────────────
    # Mirrors INGExtratransfers.dtsx: join IPS_DAILY (non-return, Whs='ING') with TUTLIV.dbo.INGQTY
    # Output rows where INGOH - IPS_sold_qty < 0
    try:
        ips_qtys = (
            daily_df[(daily_df['Otype'] != 'Return') & (daily_df['Whs'] == 'ING')]
            .groupby('ISBN')['Qty'].sum()
            .reset_index()
        )
        ips_qtys.columns = ['ISBN', 'IPS_Qty']

        with get_db() as conn:
            ingqty_df = pd.read_sql('SELECT ISBN, CAST(INGOH AS int) AS INGOH FROM TUTLIV.dbo.INGQTY', con=conn)

        ing_merged = ingqty_df.merge(ips_qtys, on='ISBN', how='inner')
        ing_merged['TOTAL_QTY_ING'] = ing_merged['INGOH'] - ing_merged['IPS_Qty']
        ing_transfers = ing_merged[ing_merged['TOTAL_QTY_ING'] < 0][['ISBN', 'TOTAL_QTY_ING']].copy()

        ing_path = os.path.join(output_path, 'ING_Transfers.csv')
        ing_transfers.to_csv(ing_path, index=False)
        logger.info(f"ING_Transfers.csv written: {len(ing_transfers)} rows → {ing_path}")
    except Exception as e:
        logger.warning(f"ING_Transfers.csv skipped — could not query TUTLIV.dbo.INGQTY: {e}")

    # ─── Steps 27-29: CDP file → LOCKEDT.TXT + INPRO.TXT ─────────────────────
    logger.info("Processing CDP file")
    cdp_df = proccess_cdp_file(CDP_path)

    # LOCKEDT.TXT — mirrors LOCKED_EXPORT.dtsx: Invcode='QH' AND San='631760X'
    # Output: F1,Fdate,San,ISBN10,F5,ISBN,Invcode,QTY
    lockedt_df = cdp_df[
        (cdp_df['Invcode'] == 'QH') & (cdp_df['San'] == '631760X')
    ][['F1', 'Fdate', 'San', 'ISBN10', 'F5', 'ISBN', 'Invcode', 'QTY']].copy()

    lockedt_path = os.path.join(output_path, 'LOCKEDT.TXT')
    lockedt_df.to_csv(lockedt_path, index=False, header=False)
    logger.info(f"LOCKEDT.TXT written: {len(lockedt_df)} rows → {lockedt_path}")

    # INPRO.TXT — mirrors Inpro_Export.dtsx: Invcode='OP' AND SAN='631760X'
    # Output: all 13 columns with header row
    inpro_df = cdp_df[
        (cdp_df['Invcode'] == 'OP') & (cdp_df['San'] == '631760X')
    ].copy()

    inpro_path = os.path.join(output_path, 'INPRO.TXT')
    inpro_df.to_csv(inpro_path, index=False, header=True)
    logger.info(f"INPRO.TXT written: {len(inpro_df)} rows → {inpro_path}")

    logger.info("Manual rerun complete — all output files generated.")
