import logging
from helpers.db_conn import get_db
from helpers.context import DailyFilesContext
import pandas as pd
from openpyxl import load_workbook as pyxl_load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font


def generate_sage_uploads():

    with get_db() as tutliv:
            logging.info("Selecting sl header")
            sl_header = pd.read_sql(
            """
                SELECT Order_id as ORDUNIQ, Ordnum as ORDNUMBER, Billto as CUSTOMER, Ponumber as PONUMBER, Pdate as ORDDATE, Rep_inv as [DESC], Traninfo as COMMENT, Post as POSTINV
                FROM IPS.dbo.IPS_DAILY
                WHERE Otype <> 'Return' 
                AND Line_num = 1 
                AND Discount <> 100.0
                ORDER BY Order_id;
            """, con = tutliv)

            sl_header['ORDDATE'] = pd.to_datetime(sl_header['ORDDATE'])
            sl_header['ORDDATE'] = sl_header['ORDDATE'].apply(lambda x: f"{x.month}/{x.day}/{x.year}")

            logging.info(f'successfully selected {sl_header.shape[0]} records for sl_header')

            logging.info("selecting sl_detail")

            sl_detail = pd.read_sql(
                """
                SELECT Order_id as ORDUNIQ, Line_num as LINENUM, ISBN as ITEM, Whs as [LOCATION], Qty as QTYORDERED, Price as PRIUNTPRC, Discount as DISCPER, Repqty as QTYSHIPPED
                FROM IPS.dbo.IPS_DAILY
                WHERE Otype <> 'Return' AND Discount <> 100.0
                ORDER BY Order_id, Line_num;
                """, con = tutliv)

            logging.info(f"Selected {sl_detail.shape[0]} rows for sl_detail")
            

            sl_filename = "SL_SAGE_UPLOAD.xlsx"

            for col in sl_header.columns:
                sl_header[col] = sl_header[col].astype(str)
            for col in sl_detail.columns:
                sl_detail[col] = sl_detail[col].astype(str)

            logging.info('writing dfs to excel sheet')
            with pd.ExcelWriter(DailyFilesContext.daily_files_path().joinpath(sl_filename)) as f:
                sl_header.to_excel(f,sheet_name = "Orders", index = False)
                sl_detail.to_excel(f,sheet_name = 'Order_Details', index = False)
            
            wb: Workbook = pyxl_load_workbook(DailyFilesContext.daily_files_path().joinpath(sl_filename))
            for sheet in ['Orders', 'Order_Details']:
                ws: Worksheet = wb[sheet]
                for col_idx in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].number_format = '@'
                for cell in ws[1]:
                    cell.font = Font(bold=False, underline='none')
                last_row = ws.max_row
                last_col = get_column_letter(ws.max_column)
                cell_range = f"'{sheet}'!$A$1:${last_col}${last_row}"
                wb.defined_names[sheet] = DefinedName(name=sheet, attr_text=cell_range)
            wb.save(DailyFilesContext.daily_files_path().joinpath(sl_filename))
            wb.close()

            logging.info("successfully wrote sl header and detail")

            logging.info('selecting credit header')
            credit_header = pd.read_sql(
                """
                SELECT Order_id as CRDUNIQ, Ordnum as ORDNUMBER, Billto as CUSTOMER, Ponumber as PONUMBER, Pdate as ORDDATE
                FROM IPS.dbo.IPS_DAILY
                WHERE Otype = 'Return' AND Line_num = 1
                ORDER BY Order_id;
                """, con = tutliv)

            credit_header['ORDDATE'] = pd.to_datetime(credit_header['ORDDATE'])
            credit_header['ORDDATE'] = credit_header['ORDDATE'].apply(lambda x: f"{x.month}/{x.day}/{x.year}")

            logging.info(f"selected {credit_header.shape[0]} credit header")
            logging.info(f"selecting credit details")
            credit_details = pd.read_sql(
                """
                SELECT Order_id as CRDUNIQ, Line_num as LINENUM, ISBN as ITEM, Whs as [LOCATION], Qty as [QTYRETURN], Price as PRIUNTPRC, Discount as DISCPER
                FROM IPS.dbo.IPS_DAILY
                WHERE Otype = 'Return'
                ORDER BY Order_id, Line_num;
                """, con = tutliv)
            
            credit_details['LOCATION'] = 'DAMAGE'
            logging.info(f"selected {credit_details.shape[0]} rows for credit detail")
            
            cr_filename = "CR_SAGE_UPLOAD.xlsx"

            for col in credit_header.columns:
                credit_header[col] = credit_header[col].astype(str)
            for col in credit_details.columns:
                credit_details[col] = credit_details[col].astype(str)

            #need to make blank dfs for multiple sheets or the upload will throw error
            credit_debit_detail_serial_nos = pd.DataFrame(columns = [
                "CRDUNIQ",
                "LINENUM",
                "SERIALNUMF",
                "DETAILNUM",
                "COST",
                "QTY"
            ])
            credit_debit_detail_lot_numbers = pd.DataFrame(
                columns = [
                    "CRDUNIQ",
                    "LINENUM",
                    "LOTNUMF",
                    "DETAILNUM",
                    "EXPIRYDATE",
                    "STKQTY",
                    "QTY",
                    "INVSTKQTY",
                    "COST"
                ]
            )

            crd_dbn_comments_instructions = pd.DataFrame(
                columns = [
                    "CRDUNIQ",
                    "UNIQUIFIER",
                    "DETAILNUM",
                    "COINTYPE",
                    "COIN"
                ]
            )

            credit_debit_note_opt_fields = pd.DataFrame(
                columns = [
                    "CRDUNIQ",
                    "OPTFIELD",
                    "VALUE",
                    "TYPE",
                    "LENGTH",
                    "DECIMALS",
                    "ALLOWNULL",
                    "VALIDATE",
                    "SWSET",
                    "VALINDEX",
                    "VALIFTEXT",
                    "VALIFMONEY",
                    "VALIFNUM",
                    "VALIFLONG",
                    "VALIFBOOL",
                    "VALIFDATE",
                    "VALIFTIME",
                    "FDESC",
                    "VDESC"
                ]
            )

            credit_debit_detail_opt_fields = pd.DataFrame(
                columns = [
                    "CRDUNIQ",
                    "LINENUM",
                    "OPTFIELD",
                    "VALUE",
                    "TYPE",
                    "LENGTH",
                    "DECIMALS",
                    "ALLOWNULL",
                    "VALIDATE",
                    "SWSET",
                    "VALINDEX",
                    "VALIFTEXT",
                    "VALIFMONEY",
                    "VALIFNUM",
                    "VALIFLONG",
                    "VALIFBOOL",
                    "VALIFDATE",
                    "VALIFTIME",
                    "FDESC",
                    "VDESC"
                ]
            )




            logging.info("writing credits to sheet")
            with pd.ExcelWriter(DailyFilesContext.daily_files_path().joinpath(cr_filename)) as f:
                credit_header.to_excel(f, sheet_name="Credit_Debit_Notes", index=False)
                credit_details.to_excel(f, sheet_name="Credit_Debit_Details", index=False)
                credit_debit_detail_serial_nos.to_excel(f, sheet_name="Credit_Debit_Detail_Serial_Nos", index=False)
                credit_debit_detail_lot_numbers.to_excel(f, sheet_name="Credit_Debit_Detail_Lot_Numbers", index=False)
                crd_dbn_comments_instructions.to_excel(f, sheet_name="Crd_Dbn_Comments_Instructions", index=False)
                credit_debit_note_opt_fields.to_excel(f, sheet_name="Credit_Debit_Note_Opt__Fields", index=False)
                credit_debit_detail_opt_fields.to_excel(f, sheet_name="Credit_Debit_Detail_Opt__Fields", index=False)

            wb: Workbook = pyxl_load_workbook(DailyFilesContext.daily_files_path().joinpath(cr_filename))
            for sheet in [
                'Credit_Debit_Notes',
                'Credit_Debit_Details',
                'Credit_Debit_Detail_Serial_Nos',
                'Credit_Debit_Detail_Lot_Numbers',
                'Crd_Dbn_Comments_Instructions',
                'Credit_Debit_Note_Opt__Fields',
                'Credit_Debit_Detail_Opt__Fields',
            ]:
                ws: Worksheet = wb[sheet]
                for col_idx in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].number_format = '@'
                for cell in ws[1]:
                    cell.font = Font(bold=False, underline='none')
                last_row = ws.max_row
                last_col = get_column_letter(ws.max_column)
                cell_range = f"'{sheet}'!$A$1:${last_col}${last_row}"
                wb.defined_names[sheet] = DefinedName(name=sheet, attr_text=cell_range)
            wb.save(DailyFilesContext.daily_files_path().joinpath(cr_filename))
            wb.close()
            logging.info("successfully wrote credits to sheet")

if __name__ == "__main__":
    generate_sage_uploads()